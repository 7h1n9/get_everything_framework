"""目标文件解析器 — 支持 .txt .csv .xlsx .json"""

import csv
import json
import os
import re
from typing import List, Optional
from urllib.parse import urlparse


DOMAIN_PATTERN = re.compile(r"^(?:[a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}$")
IP_PATTERN = re.compile(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}(?:/\d{1,2})?$")


def normalize_target(value: str) -> Optional[str]:
    """清洗单个目标字符串: 去协议/去空白/去尾点/验证格式"""
    value = (value or "").strip()
    if not value:
        return None

    # 去掉协议前缀
    if value.startswith(("http://", "https://")):
        parsed = urlparse(value)
        value = parsed.hostname or ""
        if not value:
            return None

    value = value.strip().lower().rstrip(".")

    # 合法域名 或 IP/CIDR
    if DOMAIN_PATTERN.fullmatch(value) or IP_PATTERN.fullmatch(value):
        return value

    return None


def parse_targets_file(file_path: str) -> List[str]:
    """根据扩展名自动选择解析器"""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        return _parse_csv(file_path)
    elif ext == ".xlsx":
        return _parse_xlsx(file_path)
    elif ext == ".json":
        return _parse_json(file_path)
    else:
        return _parse_txt(file_path)


def _parse_txt(file_path: str) -> List[str]:
    """逐行解析, 每行一个目标"""
    targets = []
    with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:
        for line in f:
            t = normalize_target(line)
            if t:
                targets.append(t)
    return list(dict.fromkeys(targets))


def _parse_csv(file_path: str) -> List[str]:
    """CSV 逐格解析, 自动跳过表头行"""
    targets = []
    with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            for cell in row:
                t = normalize_target(cell)
                if t:
                    targets.append(t)
    return list(dict.fromkeys(targets))


def _parse_json(file_path: str) -> List[str]:
    """
    JSON 解析, 支持两种格式:

    1. 纯数组: ["domain1.com", "1.2.3.4"]
    2. 对象数组: [{"domain":"example.com"}, {"host":"1.2.3.4"}]
    """
    targets = []
    with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError:
            return []

    items = data if isinstance(data, list) else [data]

    for item in items:
        if isinstance(item, str):
            t = normalize_target(item)
            if t:
                targets.append(t)
        elif isinstance(item, dict):
            for key in ("domain", "host", "url", "ip", "target", "hostname"):
                val = item.get(key, "")
                if isinstance(val, str):
                    t = normalize_target(val)
                    if t:
                        targets.append(t)
                        break

    return list(dict.fromkeys(targets))


def _parse_xlsx(file_path: str) -> List[str]:
    """
    Excel 解析, 遍历所有 sheet 所有单元格

    需要 pip install openpyxl
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("解析 .xlsx 需要安装 openpyxl: pip install openpyxl")

    targets = []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    t = normalize_target(cell)
                    if t:
                        targets.append(t)
    wb.close()
    return list(dict.fromkeys(targets))


def save_normalized_targets(targets: List[str], output_path: str) -> str:
    """保存为纯文本, 每行一个目标"""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        for t in targets:
            f.write(t + "\n")
    return output_path
